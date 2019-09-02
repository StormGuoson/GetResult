# coding=utf-8
import os
import sys
import time
from openpyxl import Workbook

reload(sys)
sys.setdefaultencoding('utf-8')


class Report(object):
    def __init__(self, path, device):
        self.dict_wakeup_txt = {}
        self.sys_info = {}
        self.path = path
        self.final_wakeup = {}
        self.partial_wakeup = {}
        self.device = device
        t = time.strftime('%m-%d %H-%M', time.localtime())
        self.wb_name = u'唤醒率%s %s.xlsx' % (device, t)

    def put_wakeup_txt(self, person, path):
        self.dict_wakeup_txt[person] = path

    def parser_wakeup_txt(self):
        for name, txt in self.dict_wakeup_txt.items():
            with open(txt, 'r') as f:
                for line in f.readlines():
                    self.append_partial_data(name, line[:-1])
                    if u'唤醒率' in line.decode():
                        self.final_wakeup[name] = float(line[line.rfind(':') + 1: -2])
        self.final_wakeup[u'平均值'] = sum(self.final_wakeup.values())/len(self.final_wakeup.values())

    def get_wakeup_txt(self):
        return self.dict_wakeup_txt

    def append_partial_data(self, key, value):
        if key not in self.partial_wakeup:
            self.partial_wakeup[key] = []
        self.partial_wakeup[key].append(value)

    def check_sys_info(self):
        header = 'adb -s %s shell dumpsys package' % self.device
        cmds = ['com.tencent.wecarspeech',
                'com.tencent.wecarnavi',
                'com.tencent.music',
                'com.wt.music',
                'com.wt.music.musicwidget',
                'com.tencent.wecarnews',
                'com.tencent.wecarmas',
                'com.android.launcherWT',
                'com.android.systemui',
                'com.android.cityofphantom',
                'com.tencent.wecarcontrol',
                'com.autopai.system.settings',
                'com.autopai.car.dialer',
                'com.wtcl.filemanager',
                'com.wt.airconditioner',
                'com.wutong.fota',
                'com.autopai.usercenter',
                'com.wt.vehiclesetting',
                'com.wt.drivingrecorder',
                ]
        end = '| findstr versionName'
        self.sys_info['OS-Version'] = os.popen('adb -s %s shell getprop ro.build.display.id' % self.device).readlines()[
            0]
        self.popen(header, cmds, end)
        return self.sys_info

    def popen(self, header, cmds, end=None):
        for cmd in cmds:
            c = header + ' ' + cmd + ' ' + end
            try:
                line = os.popen(c).readlines()[0][:-1].strip()
                line = line[line.find('=') + 1:]
                self.sys_info[cmd] = line
            except IndexError as e:
                # print e.message
                # print c
                pass

    def write_excel(self):
        wb = Workbook()
        self.w_version(wb)
        self.w_total(wb)
        self.w_partial(wb)
        wb.save(os.path.join(self.path, self.wb_name))

    def w_version(self, wb):
        ws = wb.active
        ws.title = u'版本信息'
        ws.column_dimensions['a'].width = 28
        ws.column_dimensions['b'].width = 25
        for k, v in self.sys_info.items():
            ws.append([k, v])

    def w_total(self, wb):
        ws = wb.create_sheet(u'汇总')
        ws.column_dimensions['a'].width = 13
        ws.append([u'人员', u'唤醒率'])
        data = ''
        for k, v in self.final_wakeup.items():
            if k == u'平均值':
                data = v
            else:
                ws.append([k, v])
        ws.append([u'平均值', data])

    def w_partial(self, wb):
        for name in self.partial_wakeup.keys():
            ws = wb.create_sheet(name)
            for values in self.partial_wakeup[name]:
                ws.append([values])


# if __name__ == '__main__':
    # r = Report('d:\\wakeup', '123456789')
    # r.put_wakeup_txt('liuxiu', r'D:\0830-123456789\female_liuxiu\wakeup_result.txt')
    # r.put_wakeup_txt('zhaoxiangna', r'D:\0830-123456789\female_zhaoxiangna\wakeup_result.txt')
    # r.put_wakeup_txt('guoyuqiang', r'D:\0830-123456789\male_guoyuqiang\wakeup_result.txt')
    # r.put_wakeup_txt('tiankaikun', r'D:\0830-123456789\male_tiankaikun\wakeup_result.txt')
    # r.put_wakeup_txt('wangjiang', r'D:\0830-123456789\male_wangjiang\wakeup_result.txt')
    # r.parser_wakeup_txt()
    # r.check_sys_info()
    # r.write_excel()
